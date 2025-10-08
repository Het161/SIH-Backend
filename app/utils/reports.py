from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io
from datetime import datetime

def generate_user_productivity_pdf(user_data: dict) -> bytes:
    """Generate PDF report for user productivity"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"<b>Productivity Report: {user_data['user_name']}</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Date
    date_text = Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal'])
    elements.append(date_text)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary table
    data = [
        ['Metric', 'Value'],
        ['Productivity Score', f"{user_data['score']}%"],
        ['Total Tasks', str(user_data['total_tasks'])],
        ['Completed Tasks', str(user_data['completed_tasks'])],
        ['On-Time Tasks', str(user_data['on_time_tasks'])],
        ['Overdue Tasks', str(user_data['overdue_tasks'])],
        ['Avg Completion Time', f"{user_data['avg_completion_time']} hours"],
        ['Completion Rate', f"{user_data['completion_rate']}%"]
    ]
    
    table = Table(data, colWidths=[3*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return buffer.getvalue()


def generate_user_productivity_excel(user_data: dict) -> bytes:
    """Generate Excel report for user productivity"""
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productivity Report"
    
    # Header
    ws['A1'] = f"Productivity Report: {user_data['user_name']}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Headers
    ws['A4'] = "Metric"
    ws['B4'] = "Value"
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)
    
    # Data
    metrics = [
        ("Productivity Score", f"{user_data['score']}%"),
        ("Total Tasks", user_data['total_tasks']),
        ("Completed Tasks", user_data['completed_tasks']),
        ("On-Time Tasks", user_data['on_time_tasks']),
        ("Overdue Tasks", user_data['overdue_tasks']),
        ("Avg Completion Time (hours)", user_data['avg_completion_time']),
        ("Completion Rate", f"{user_data['completion_rate']}%")
    ]
    
    for idx, (metric, value) in enumerate(metrics, start=5):
        ws[f'A{idx}'] = metric
        ws[f'B{idx}'] = value
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_team_productivity_excel(team_data: dict, users_data: list) -> bytes:
    """Generate Excel report for team productivity"""
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Team Report"
    
    # Header
    ws['A1'] = f"Team Productivity Report: {team_data['department']}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Team Summary
    ws['A4'] = "Team Summary"
    ws['A4'].font = Font(size=14, bold=True)
    ws['A5'] = "Total Employees"
    ws['B5'] = team_data['total_employees']
    ws['A6'] = "Average Score"
    ws['B6'] = f"{team_data['avg_score']}%"
    ws['A7'] = "Total Tasks"
    ws['B7'] = team_data['total_tasks']
    ws['A8'] = "Completed Tasks"
    ws['B8'] = team_data['completed_tasks']
    
    # Individual employees
    ws['A10'] = "Individual Employee Metrics"
    ws['A10'].font = Font(size=14, bold=True)
    
    headers = ['Employee', 'Score', 'Total Tasks', 'Completed', 'On-Time', 'Overdue']
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=idx)
        cell.value = header
        cell.font = Font(bold=True)
    
    for idx, user in enumerate(users_data, start=12):
        ws.cell(row=idx, column=1, value=user['user_name'])
        ws.cell(row=idx, column=2, value=f"{user['score']}%")
        ws.cell(row=idx, column=3, value=user['total_tasks'])
        ws.cell(row=idx, column=4, value=user['completed_tasks'])
        ws.cell(row=idx, column=5, value=user['on_time_tasks'])
        ws.cell(row=idx, column=6, value=user['overdue_tasks'])
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
